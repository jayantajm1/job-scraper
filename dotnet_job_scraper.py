"""
.NET Job Scraper — Apify API → Excel
======================================
Scrapes 100 .NET jobs from LinkedIn, Indeed, and Glassdoor via Apify,
then saves results to a beautifully formatted Excel (.xlsx) file.

SETUP:
    pip install requests openpyxl

USAGE:
    1. Replace APIFY_API_TOKEN below with your token
       → Get it free at: https://console.apify.com/account/integrations
    2. Adjust SEARCH_LOCATION if needed
    3. Run:  python dotnet_job_scraper.py
"""

import requests
import time
import sys
import csv
import json
import os
from datetime import datetime

OPENPYXL_AVAILABLE = True
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    OPENPYXL_AVAILABLE = False

# Import application modules
try:
    from profile import CandidateProfile
    from cover_letter_gen import CoverLetterGenerator
    from application_tracker import ApplicationTracker
    from apply_engine import JobApplicationEngine
    APPLY_AVAILABLE = True
except ImportError:
    APPLY_AVAILABLE = False

# ══════════════════════════════════════════════════════════════
#  ✏️  CONFIGURATION — Set these in environment or .env file
# ══════════════════════════════════════════════════════════════

import os
from dotenv import load_dotenv

# Load environment variables from .env file if it exists
load_dotenv()

APIFY_API_TOKEN  = os.getenv('APIFY_API_TOKEN')  # Set in .env or export APIFY_API_TOKEN=your_token
if not APIFY_API_TOKEN:
    print("[ERROR] APIFY_API_TOKEN not found. Set it in .env file or environment variable.")
    print("[INFO] Create a .env file with: APIFY_API_TOKEN=your_token_here")
    sys.exit(1)
SEARCH_KEYWORD   = ".NET Developer"              # Change to "C# Developer", "ASP.NET", etc.
SEARCH_LOCATION  = "India"                       # e.g. "Kolkata", "Remote", "United States"
TOTAL_JOBS       = 100                           # Total jobs to collect
OUTPUT_DIR       = os.getenv("JOB_OUTPUT_DIR", ".")
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_FILE      = os.path.join(OUTPUT_DIR, f"dotnet_jobs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

# ── Apify Actor IDs (available on Apify Store) ──────────────
# LinkedIn Jobs Scraper  → https://apify.com/curious_coder/linkedin-jobs-scraper
# Indeed Scraper         → https://apify.com/misceres/indeed-scraper
# Glassdoor Scraper      → https://apify.com/misceres/glassdoor-scraper

SOURCES = [
    {
        "name":     "LinkedIn",
        "actor_id": "curious_coder/linkedin-jobs-scraper",
        "count":    40,
        "payload":  lambda kw, loc, n: {
            "title":    kw,
            "location": loc,
            "rows":     n,
        },
    },
    {
        "name":     "Indeed",
        "actor_id": "misceres/indeed-scraper",
        "count":    40,
        "payload":  lambda kw, loc, n: {
            "position": kw,
            "country":  "IN" if "india" in loc.lower() else "US",
            "location": loc,
            "maxItems": n,
        },
    },
    {
        "name":     "Glassdoor",
        "actor_id": "misceres/glassdoor-scraper",
        "count":    20,
        "payload":  lambda kw, loc, n: {
            "keyword":  kw,
            "location": loc,
            "maxItems": n,
            "type":     "JOBS",
        },
    },
]

APIFY_BASE = "https://api.apify.com/v2"

if not OPENPYXL_AVAILABLE:
    OUTPUT_FILE = OUTPUT_FILE.replace(".xlsx", ".csv")

# ══════════════════════════════════════════════════════════════
#  APIFY RUNNER
# ══════════════════════════════════════════════════════════════

def run_actor(actor_id: str, payload: dict, max_items: int) -> list[dict]:
    """Start an Apify actor run, wait for completion, and return dataset items."""
    params  = {"token": APIFY_API_TOKEN}
    headers = {"Content-Type": "application/json"}
    actor_api_id = actor_id.replace("/", "~")

    # Start run
    resp = requests.post(
        f"{APIFY_BASE}/acts/{actor_api_id}/runs",
        json=payload, headers=headers, params=params, timeout=30
    )
    if resp.status_code == 401:
        print("  ❌ Invalid Apify API token. Check your APIFY_API_TOKEN.")
        sys.exit(1)
    resp.raise_for_status()

    run_id = resp.json()["data"]["id"]
    print(f"     Run ID: {run_id} — waiting for results...")

    # Poll until done (max 6 minutes)
    status_url = f"{APIFY_BASE}/actor-runs/{run_id}"
    for i in range(72):
        time.sleep(5)
        s = requests.get(status_url, params=params, timeout=15).json()["data"]["status"]
        print(f"     [{i*5}s] Status: {s}", end="\r")
        if s == "SUCCEEDED":
            print()
            break
        if s in ("FAILED", "ABORTED", "TIMED-OUT"):
            print(f"\n  ⚠️  Actor ended with status: {s}. Skipping this source.")
            return []
    else:
        print(f"\n  ⚠️  Timeout waiting for actor {actor_id}. Skipping.")
        return []

    # Fetch items
    dataset_id = requests.get(status_url, params=params, timeout=15).json()["data"]["defaultDatasetId"]
    items = requests.get(
        f"{APIFY_BASE}/datasets/{dataset_id}/items",
        params={**params, "limit": max_items}, timeout=30
    ).json()
    return items if isinstance(items, list) else []


# ══════════════════════════════════════════════════════════════
#  DATA NORMALIZERS — convert each source's raw JSON → standard dict
# ══════════════════════════════════════════════════════════════

def normalize_linkedin(item: dict, idx: int) -> dict:
    return {
        "No.":        idx,
        "Source":     "LinkedIn",
        "Job Title":  item.get("title") or item.get("jobTitle", "N/A"),
        "Company":    item.get("companyName") or item.get("company", "N/A"),
        "Location":   item.get("location", "N/A"),
        "Job Type":   item.get("employmentType") or item.get("jobType", "N/A"),
        "Experience": item.get("seniorityLevel") or item.get("experienceLevel", "N/A"),
        "Salary":     item.get("salary") or item.get("salaryRange", "N/A"),
        "Posted Date":item.get("postedAt") or item.get("publishedAt", "N/A"),
        "Apply Link": item.get("jobUrl") or item.get("url") or item.get("link", "N/A"),
        "Description":_trim(item.get("description") or item.get("descriptionText", "")),
    }

def normalize_indeed(item: dict, idx: int) -> dict:
    return {
        "No.":        idx,
        "Source":     "Indeed",
        "Job Title":  item.get("positionName") or item.get("title", "N/A"),
        "Company":    item.get("company", "N/A"),
        "Location":   item.get("location", "N/A"),
        "Job Type":   item.get("jobType") or item.get("employmentType", "N/A"),
        "Experience": item.get("experienceLevel", "N/A"),
        "Salary":     item.get("salary", "N/A"),
        "Posted Date":item.get("postedAt") or item.get("date", "N/A"),
        "Apply Link": item.get("url") or item.get("externalApplyLink") or item.get("applyUrl", "N/A"),
        "Description":_trim(item.get("description", "")),
    }

def normalize_glassdoor(item: dict, idx: int) -> dict:
    return {
        "No.":        idx,
        "Source":     "Glassdoor",
        "Job Title":  item.get("jobTitle") or item.get("title", "N/A"),
        "Company":    item.get("employerName") or item.get("company", "N/A"),
        "Location":   item.get("location") or item.get("locationName", "N/A"),
        "Job Type":   item.get("jobType", "N/A"),
        "Experience": item.get("seniorityLevel", "N/A"),
        "Salary":     item.get("salaryEstimate") or item.get("salary", "N/A"),
        "Posted Date":item.get("discoveryDate") or item.get("postedDate", "N/A"),
        "Apply Link": item.get("jobListingUrl") or item.get("url", "N/A"),
        "Description":_trim(item.get("jobDescription") or item.get("description", "")),
    }

NORMALIZERS = {
    "LinkedIn":  normalize_linkedin,
    "Indeed":    normalize_indeed,
    "Glassdoor": normalize_glassdoor,
}

def _trim(text: str, max_len: int = 500) -> str:
    text = (text or "").strip().replace("\n", " ")
    return text[:max_len] + "..." if len(text) > max_len else text


def _excel_safe(value):
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return value


def write_csv(jobs: list[dict], filepath: str):
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(jobs)
    print(f"\n✅ CSV file saved: {filepath}")
    print(f"   • Total rows: {len(jobs)}")


# ══════════════════════════════════════════════════════════════
#  EXCEL WRITER
# ══════════════════════════════════════════════════════════════

# Brand colors
COLOR_HEADER_BG  = "1F4E79"   # Dark blue header
COLOR_HEADER_FG  = "FFFFFF"   # White text
COLOR_LINKEDIN   = "DDEEFF"   # Light blue
COLOR_INDEED     = "DFFDE0"   # Light green
COLOR_GLASSDOOR  = "FFF3CC"   # Light yellow
COLOR_ALT_ROW    = "F5F5F5"   # Zebra stripe
SOURCE_COLORS    = {"LinkedIn": COLOR_LINKEDIN, "Indeed": COLOR_INDEED, "Glassdoor": COLOR_GLASSDOOR}

COLUMNS = ["No.", "Source", "Job Title", "Company", "Location",
           "Job Type", "Experience", "Salary", "Posted Date", "Apply Link", "Description",
           "Apply Status", "Applied DateTime", "Resume Version", "Cover Letter", 
           "Application Method", "Recruiter", "Follow-up Date"]
COL_WIDTHS = [5, 12, 30, 25, 20, 15, 18, 20, 15, 45, 60, 15, 18, 18, 15, 18, 15, 15]

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_excel(jobs: list[dict], filepath: str):
    wb = Workbook()

    # ── Sheet 1: All Jobs ──────────────────────────────────────
    ws = wb.active
    ws.title = "All Jobs"
    ws.freeze_panes = "A2"

    # Title banner
    ws.merge_cells("A1:K1")
    ws["A1"] = f"🔍 .NET Job Listings  |  Keyword: {SEARCH_KEYWORD}  |  Location: {SEARCH_LOCATION}  |  Scraped: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=12, color=COLOR_HEADER_FG)
    ws["A1"].fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers (row 2)
    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FG)
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, job in enumerate(jobs, start=3):
        src_color = SOURCE_COLORS.get(job.get("Source", ""), COLOR_ALT_ROW)
        fill = PatternFill("solid", fgColor=src_color if row_idx % 2 == 0 else "FFFFFF")
        for col_idx, col in enumerate(COLUMNS, start=1):
            val  = _excel_safe(job.get(col, ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(col == "Description"))
            cell.border    = thin_border()
            cell.fill      = fill
            # Make Apply Link a hyperlink
            if col == "Apply Link" and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
        ws.row_dimensions[row_idx].height = 60 if any(len(str(job.get(c, ""))) > 100 for c in ["Description"]) else 30

    # Auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(jobs) + 2}"

    # ── Sheet 2: Summary Dashboard ─────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 35
    ws2.column_dimensions["D"].width = 20

    # Summary header
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "📊 Scraping Summary"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FG)
    ws2["A1"].fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    summary_data = [
        ("Search Keyword",   SEARCH_KEYWORD,  "", ""),
        ("Location",         SEARCH_LOCATION, "", ""),
        ("Scrape Date",      datetime.now().strftime("%d %b %Y %H:%M"), "", ""),
        ("Total Jobs Found", len(jobs), "", ""),
        ("", "", "", ""),
        ("Source Breakdown", "Count", "% of Total", ""),
    ]

    source_counts = {}
    for j in jobs:
        s = j.get("Source", "Unknown")
        source_counts[s] = source_counts.get(s, 0) + 1

    for src, cnt in source_counts.items():
        pct = f"{cnt/len(jobs)*100:.1f}%" if jobs else "0%"
        summary_data.append((src, cnt, pct, ""))

    summary_data += [
        ("", "", "", ""),
        ("Top Companies", "Jobs Count", "", ""),
    ]
    company_counts = {}
    for j in jobs:
        c = j.get("Company", "Unknown")
        company_counts[c] = company_counts.get(c, 0) + 1
    for company, cnt in sorted(company_counts.items(), key=lambda x: -x[1])[:10]:
        summary_data.append((company, cnt, "", ""))

    for r_idx, row in enumerate(summary_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = Font(name="Arial", size=10, bold=(r_idx in [6, 12]))
            cell.alignment = Alignment(vertical="center")
            cell.border    = thin_border()
            if r_idx == 6 or r_idx == 12:
                cell.fill = PatternFill("solid", fgColor="BDD7EE")
                cell.font = Font(name="Arial", bold=True, size=10)

    # ── Sheet 3: Per-Source sheets ─────────────────────────────
    for source in ["LinkedIn", "Indeed", "Glassdoor"]:
        src_jobs = [j for j in jobs if j.get("Source") == source]
        if not src_jobs:
            continue
        ws_src = wb.create_sheet(source)
        ws_src.freeze_panes = "A2"
        color = SOURCE_COLORS.get(source, "FFFFFF")

        for col_idx, (col_name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
            cell = ws_src.cell(row=1, column=col_idx, value=col_name)
            cell.font      = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FG)
            cell.fill      = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()
            ws_src.column_dimensions[get_column_letter(col_idx)].width = width

        for r_idx, job in enumerate(src_jobs, start=2):
            fill = PatternFill("solid", fgColor=color if r_idx % 2 == 0 else "FFFFFF")
            for c_idx, col in enumerate(COLUMNS, start=1):
                val  = _excel_safe(job.get(col, ""))
                cell = ws_src.cell(row=r_idx, column=c_idx, value=val)
                cell.font      = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=(col == "Description"))
                cell.border    = thin_border()
                cell.fill      = fill
                if col == "Apply Link" and val and str(val).startswith("http"):
                    cell.hyperlink = val
                    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            ws_src.row_dimensions[r_idx].height = 30

        ws_src.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(src_jobs) + 1}"

    wb.save(filepath)
    print(f"\n✅ Excel file saved: {filepath}")
    print(f"   • Sheets: 'All Jobs', 'Summary', + one sheet per source")
    print(f"   • Total rows: {len(jobs)}")


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

def main():
    if APIFY_API_TOKEN == "YOUR_APIFY_API_TOKEN_HERE":
        print("❌ ERROR: Please set your APIFY_API_TOKEN in the script before running.")
        print("   Get a free token at: https://console.apify.com/account/integrations")
        sys.exit(1)

    print("=" * 60)
    print(f"  .NET Job Scraper + Application Assistant")
    print(f"  Keyword : {SEARCH_KEYWORD}")
    print(f"  Location: {SEARCH_LOCATION}")
    print(f"  Target  : {TOTAL_JOBS} jobs from {len(SOURCES)} sources")
    print("=" * 60)

    all_jobs = []
    job_counter = 1

    for source in SOURCES:
        name     = source["name"]
        actor_id = source["actor_id"]
        count    = source["count"]
        payload  = source["payload"](SEARCH_KEYWORD, SEARCH_LOCATION, count)

        print(f"\n🔍 [{name}] Starting scrape (target: {count} jobs)...")
        try:
            raw_items = run_actor(actor_id, payload, count)
        except Exception as e:
            print(f"  ⚠️  Failed to scrape {name}: {e}")
            raw_items = []

        normalize = NORMALIZERS[name]
        for item in raw_items:
            try:
                job = normalize(item, job_counter)
                all_jobs.append(job)
                job_counter += 1
                if job_counter % 10 == 0:
                    print(f"  📌 Collected {job_counter - 1} jobs so far...")
            except Exception:
                pass  # Skip malformed items

        print(f"  ✅ {name}: {len([j for j in all_jobs if j['Source'] == name])} jobs collected")

    if not all_jobs:
        print("\n❌ No jobs were collected. Check your API token and internet connection.")
        sys.exit(1)

    print(f"\n📊 Total jobs collected: {len(all_jobs)}")
    if OPENPYXL_AVAILABLE:
        print(f"💾 Writing to Excel: {OUTPUT_FILE}")
        write_excel(all_jobs, OUTPUT_FILE)
    else:
        print("⚠️  openpyxl is not installed; exporting CSV instead of Excel.")
        print(f"💾 Writing to CSV: {OUTPUT_FILE}")
        write_csv(all_jobs, OUTPUT_FILE)

    print("\n🎉 Job scraping completed!")
    print(f"   File: {OUTPUT_FILE}")
    
    # Offer application assistant
    if APPLY_AVAILABLE and OPENPYXL_AVAILABLE:
        print("\n" + "="*60)
        app_mode = input("Application Mode? (interactive/auto/skip): ").strip().lower()
        
        if app_mode in ['interactive', 'i']:
            try:
                profile = CandidateProfile("mydetail.md")
                print(profile.get_profile_summary())
                
                engine = JobApplicationEngine(profile, OUTPUT_FILE)
                
                # Ask how many jobs to apply to
                num_jobs = input("\nHow many jobs would you like to apply to? (leave blank for all): ").strip()
                max_apps = int(num_jobs) if num_jobs.isdigit() else None
                
                engine.run_batch_applications(all_jobs, max_applications=max_apps)
                
            except FileNotFoundError:
                print("[ERROR] mydetail.md not found. Please create your profile first.")
            except Exception as e:
                print(f"[ERROR] Error in application mode: {e}")
        
        elif app_mode in ['auto', 'a']:
            try:
                profile = CandidateProfile("mydetail.md")
                print(profile.get_profile_summary())
                
                engine = JobApplicationEngine(profile, OUTPUT_FILE)
                
                # Ask how many jobs to auto-apply to
                num_jobs = input("\nHow many jobs would you like to auto-apply to? (leave blank for all): ").strip()
                max_apps = int(num_jobs) if num_jobs.isdigit() else None
                
                # Confirm before auto-apply
                confirm = input("\n[WARNING] Auto-apply will apply without review. Continue? (yes/no): ").strip().lower()
                if confirm in ['yes', 'y']:
                    engine.run_auto_applications(all_jobs, max_applications=max_apps)
                else:
                    print("[INFO] Auto-apply cancelled.")
                
            except FileNotFoundError:
                print("[ERROR] mydetail.md not found. Please create your profile first.")
            except Exception as e:
                print(f"[ERROR] Error in auto-apply mode: {e}")
        
        else:
            print("[INFO] Skipping application mode. You can apply to jobs manually later.")
    else:
        if not APPLY_AVAILABLE:
            print("\n[NOTE] Application modules not available. Scraping complete.")
        if not OPENPYXL_AVAILABLE:
            print("[NOTE] openpyxl required for application tracking. Install with: pip install openpyxl")


if __name__ == "__main__":
    main()
