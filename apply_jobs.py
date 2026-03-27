#!/usr/bin/env python
"""
Standalone Job Application Manager
===================================
Use this script to apply to jobs from an existing Excel file
without needing to re-scrape.

Usage:
    python apply_jobs.py <excel_file> [mode] [max_applications]

Modes:
    interactive (or i) - Review each job before applying (default)
    auto (or a)        - Auto-apply without review

Examples:
    python apply_jobs.py dotnet_jobs_20260327_084640.xlsx interactive 10
    python apply_jobs.py dotnet_jobs_20260327_084640.xlsx auto 20
"""

import sys
import os
from openpyxl import load_workbook
from profile import CandidateProfile
from apply_engine import JobApplicationEngine


def load_jobs_from_excel(excel_file: str) -> list:
    """Load job data from Excel file."""
    if not os.path.exists(excel_file):
        print(f"[ERROR] File not found: {excel_file}")
        return []
    
    jobs = []
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Get column headers (assuming row 2 has headers)
        headers = [cell.value for cell in ws[2]]
        
        # Load data rows (starting from row 3)
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            job_dict = {headers[i]: row[i] for i in range(len(headers))}
            if job_dict.get("Job Title"):  # Skip empty rows
                jobs.append(job_dict)
        
        print(f"[OK] Loaded {len(jobs)} jobs from {excel_file}\n")
        return jobs
    
    except Exception as e:
        print(f"[ERROR] Error loading jobs: {e}")
        return []


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    
    excel_file = sys.argv[1]
    mode = sys.argv[2].lower() if len(sys.argv) > 2 else "interactive"
    max_apps = int(sys.argv[3]) if len(sys.argv) > 3 else None
    
    # Load jobs from Excel
    all_jobs = load_jobs_from_excel(excel_file)
    if not all_jobs:
        sys.exit(1)
    
    # Load candidate profile
    try:
        profile = CandidateProfile("mydetail.md")
        print(profile.get_profile_summary())
    except FileNotFoundError:
        print("[ERROR] mydetail.md not found. Create your profile first.")
        sys.exit(1)
    
    # Initialize engine
    engine = JobApplicationEngine(profile, excel_file)
    
    # Run in selected mode
    if mode in ['auto', 'a']:
        print("[INFO] Running in AUTO-APPLY mode (no interaction)\n")
        engine.run_auto_applications(all_jobs, max_applications=max_apps)
    else:
        print("[INFO] Running in INTERACTIVE mode (review each job)\n")
        engine.run_batch_applications(all_jobs, max_applications=max_apps)


if __name__ == "__main__":
    main()
