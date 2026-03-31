"""Flask API wrapper for the job scraper command-line workflows."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import logging
import os
import subprocess
import sys
import threading
import time
import uuid

from dotenv import load_dotenv
from flask import Flask, jsonify, request
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load environment variables from .env file
load_dotenv()

# Ensure UTF-8 encoding for subprocess outputs (fixes emoji/unicode issues on Windows)
os.environ["PYTHONIOENCODING"] = "utf-8"

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("JOB_DATA_DIR", str(BASE_DIR))).resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)
TASKS: dict[str, dict] = {}
TASKS_LOCK = threading.Lock()
OUTPUT_CHAR_LIMIT = 20000

ALLOWED_ORIGINS_RAW = os.getenv("ALLOWED_ORIGINS", "*")
ALLOWED_ORIGINS = {
    item.strip()
    for item in ALLOWED_ORIGINS_RAW.split(",")
    if item.strip()
}
RATE_LIMIT_WINDOW_SECONDS = int(os.getenv("RATE_LIMIT_WINDOW_SECONDS", "60"))
RATE_LIMIT_MAX_REQUESTS = int(os.getenv("RATE_LIMIT_MAX_REQUESTS", "120"))
RATE_LIMIT_STATE: dict[str, list[float]] = {}
RATE_LIMIT_LOCK = threading.Lock()

logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
LOGGER = logging.getLogger("api_server")

app = Flask(__name__)


def _is_origin_allowed(origin: str | None) -> bool:
    if not origin:
        return True
    if "*" in ALLOWED_ORIGINS:
        return True
    return origin in ALLOWED_ORIGINS


def _is_rate_limited(client_key: str) -> bool:
    now = time.time()
    cutoff = now - RATE_LIMIT_WINDOW_SECONDS

    with RATE_LIMIT_LOCK:
        existing = RATE_LIMIT_STATE.get(client_key, [])
        recent = [stamp for stamp in existing if stamp >= cutoff]
        if len(recent) >= RATE_LIMIT_MAX_REQUESTS:
            RATE_LIMIT_STATE[client_key] = recent
            return True

        recent.append(now)
        RATE_LIMIT_STATE[client_key] = recent
        return False


@app.before_request
def before_request_checks():
    origin = request.headers.get("Origin")
    if not _is_origin_allowed(origin):
        return jsonify({"error": "Origin not allowed"}), 403

    if request.path.startswith("/api/"):
        client_ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown")
        if _is_rate_limited(client_ip):
            return jsonify({"error": "Too many requests"}), 429


@app.after_request
def add_cors_headers(response):
    origin = request.headers.get("Origin")
    if origin and _is_origin_allowed(origin):
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Vary"] = "Origin"
    elif "*" in ALLOWED_ORIGINS:
        response.headers["Access-Control-Allow-Origin"] = "*"

    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    return response


def _tail_text(value: str, limit: int = OUTPUT_CHAR_LIMIT) -> str:
    if len(value) <= limit:
        return value
    return value[-limit:]


def _serialize_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _resolve_excel_file(filename: str) -> Path:
    if not filename:
        raise ValueError("Missing file name")

    candidate = (DATA_DIR / filename).resolve()
    if DATA_DIR not in candidate.parents:
        raise ValueError("Invalid file path")
    if candidate.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported")
    if not candidate.exists():
        raise FileNotFoundError(f"File not found: {filename}")
    return candidate


def _list_excel_files() -> list[dict]:
    files = sorted(
        DATA_DIR.glob("dotnet_jobs_*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return [
        {
            "name": file.name,
            "sizeBytes": file.stat().st_size,
            "updatedAt": datetime.fromtimestamp(file.stat().st_mtime).isoformat(),
        }
        for file in files
    ]


def _load_jobs_from_excel(excel_path: Path, limit: int = 100) -> list[dict]:
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook.active

    headers = [cell.value for cell in worksheet[2]]
    jobs = []

    for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, values_only=True):
        row_data = {headers[idx]: _serialize_value(row[idx]) for idx in range(len(headers))}
        if row_data.get("Job Title"):
            jobs.append(row_data)
        if len(jobs) >= limit:
            break

    return jobs


def _init_tracking_columns(excel_path: Path) -> bool:
    """Initialize application tracking columns if they don't exist."""
    try:
        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        
        headers = [cell.value for cell in worksheet[2]]
        
        tracking_columns = [
            "Apply Status", 
            "Applied DateTime", 
            "Resume Version",
            "Cover Letter", 
            "Application Method", 
            "Recruiter", 
            "Follow-up Date"
        ]
        
        # Find the starting column for new tracking columns
        start_col = len(headers) + 1
        
        # Add tracking columns if they don't exist
        for idx, col_name in enumerate(tracking_columns):
            if col_name not in headers:
                col_letter = chr(64 + start_col + idx) if start_col + idx <= 26 else ""
                cell = worksheet.cell(row=2, column=start_col + idx, value=col_name)
                cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        workbook.save(excel_path)
        return True
    except Exception as e:
        print(f"Error initializing tracking columns: {e}")
        return False


def _mark_job_applied(excel_path: Path, job_no: int | str) -> bool:
    """Mark a job as applied with current timestamp."""
    try:
        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        
        # First ensure tracking columns exist
        _init_tracking_columns(excel_path)
        
        # Reload to get updated headers
        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        
        headers = [cell.value for cell in worksheet[2]]
        
        # Find row with matching job No.
        target_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet.max_row), start=3):
            if row[0].value == job_no:
                target_row = row_idx
                break
        
        if not target_row:
            return False
        
        # Find column indices for tracking columns
        status_col = None
        datetime_col = None
        
        for idx, header in enumerate(headers):
            if header == "Apply Status":
                status_col = idx + 1
            elif header == "Applied DateTime":
                datetime_col = idx + 1
        
        # Update cells
        if status_col:
            cell = worksheet.cell(row=target_row, column=status_col)
            cell.value = "Applied"
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        if datetime_col:
            worksheet.cell(row=target_row, column=datetime_col).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        workbook.save(excel_path)
        return True
    except Exception as e:
        print(f"Error marking job applied: {e}")
        return False


def _mark_all_jobs_applied(excel_path: Path) -> dict:
    """Mark ALL jobs in the file as applied with current timestamp."""
    try:
        # First ensure tracking columns exist
        _init_tracking_columns(excel_path)
        
        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        
        headers = [cell.value for cell in worksheet[2]]
        
        # Find column indices for tracking columns
        status_col = None
        datetime_col = None
        
        for idx, header in enumerate(headers):
            if header == "Apply Status":
                status_col = idx + 1
            elif header == "Applied DateTime":
                datetime_col = idx + 1
        
        applied_count = 0
        timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Mark all job rows as applied
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet.max_row), start=3):
            if row[0].value is not None and row[0].value != "":  # Skip empty rows
                if status_col:
                    cell = worksheet.cell(row=row_idx, column=status_col)
                    cell.value = "Applied"
                    cell.fill = green_fill
                
                if datetime_col:
                    worksheet.cell(row=row_idx, column=datetime_col).value = timestamp_str
                
                applied_count += 1
        
        workbook.save(excel_path)
        
        return {
            "success": True,
            "appliedCount": applied_count,
            "timestamp": timestamp_str,
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "appliedCount": 0,
        }


def _run_subprocess_task(task_id: str, command: list[str], extra_env: dict[str, str] | None = None) -> None:
    started = time.time()
    output = ""
    return_code = None

    try:
        process = subprocess.Popen(
            command,
            cwd=str(BASE_DIR),
            env={**os.environ.copy(), **(extra_env or {})},
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        std_output, _ = process.communicate()
        output = std_output or ""
        return_code = process.returncode
    except Exception as exc:  # pragma: no cover - runtime protection
        output = f"Failed to run command: {exc}"
        return_code = -1

    finished = time.time()
    status = "completed" if return_code == 0 else "failed"

    with TASKS_LOCK:
        task = TASKS[task_id]
        task["status"] = status
        task["returnCode"] = return_code
        task["finishedAt"] = datetime.now().isoformat()
        task["durationSeconds"] = round(finished - started, 2)
        task["output"] = _tail_text(output)


def _start_task(name: str, command: list[str], extra_env: dict[str, str] | None = None) -> dict:
    task_id = str(uuid.uuid4())
    task = {
        "id": task_id,
        "name": name,
        "command": command,
        "status": "running",
        "returnCode": None,
        "startedAt": datetime.now().isoformat(),
        "finishedAt": None,
        "durationSeconds": None,
        "output": "",
    }

    with TASKS_LOCK:
        TASKS[task_id] = task

    thread = threading.Thread(target=_run_subprocess_task, args=(task_id, command, extra_env), daemon=True)
    thread.start()

    return task


@app.get("/api/health")
def health_check():
    with TASKS_LOCK:
        running = sum(1 for task in TASKS.values() if task["status"] == "running")

    return jsonify(
        {
            "ok": True,
            "runningTasks": running,
            "timestamp": datetime.now().isoformat(),
            "environment": os.getenv("FLASK_ENV", "development"),
        }
    )


@app.route("/api/<path:subpath>", methods=["OPTIONS"])
def preflight_options(subpath: str):
    return ("", 204)


@app.get("/api/files")
def list_files():
    files = _list_excel_files()
    return jsonify({"files": files, "count": len(files)})


@app.get("/api/jobs")
def list_jobs():
    filename = request.args.get("file", "")
    limit = int(request.args.get("limit", "120"))

    try:
        excel_file = _resolve_excel_file(filename)
        jobs = _load_jobs_from_excel(excel_file, max(1, min(limit, 500)))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify({"file": excel_file.name, "count": len(jobs), "jobs": jobs})


@app.get("/api/tasks")
def list_tasks():
    with TASKS_LOCK:
        items = list(TASKS.values())
    items.sort(key=lambda item: item["startedAt"], reverse=True)
    return jsonify({"tasks": items[:25], "count": len(items)})


@app.get("/api/tasks/<task_id>")
def get_task(task_id: str):
    with TASKS_LOCK:
        task = TASKS.get(task_id)
    if not task:
        return jsonify({"error": "Task not found"}), 404
    return jsonify(task)


@app.post("/api/scrape")
def start_scrape():
    command = [sys.executable, "dotnet_job_scraper.py"]
    task = _start_task("scrape", command, {"JOB_OUTPUT_DIR": str(DATA_DIR)})
    return jsonify(task), 202


@app.post("/api/auto-apply")
def start_auto_apply():
    payload = request.get_json(silent=True) or {}
    filename = payload.get("file", "")
    max_applications = payload.get("maxApplications")

    try:
        excel_file = _resolve_excel_file(filename)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400

    command = [sys.executable, "auto_apply.py", str(excel_file)]
    if max_applications:
        command.append(str(max_applications))

    task = _start_task("auto-apply", command)
    return jsonify(task), 202


@app.post("/api/apply")
def start_apply():
    payload = request.get_json(silent=True) or {}
    filename = payload.get("file", "")
    mode = payload.get("mode", "interactive")
    max_applications = payload.get("maxApplications")

    if mode not in {"interactive", "auto", "i", "a"}:
        return jsonify({"error": "Mode must be interactive or auto"}), 400

    try:
        excel_file = _resolve_excel_file(filename)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400

    command = [sys.executable, "apply_jobs.py", str(excel_file), mode]
    if max_applications:
        command.append(str(max_applications))

    task = _start_task("apply", command)
    return jsonify(task), 202


@app.post("/api/init-tracking")
def init_tracking():
    """Initialize tracking columns for a specific Excel file."""
    payload = request.get_json(silent=True) or {}
    filename = payload.get("file", "")
    
    try:
        excel_file = _resolve_excel_file(filename)
        success = _init_tracking_columns(excel_file)
        if success:
            return jsonify({"status": "ok", "message": "Tracking columns initialized", "file": filename}), 200
        else:
            return jsonify({"error": "Failed to initialize tracking columns"}), 500
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/mark-applied")
def mark_applied():
    """Mark a specific job as applied."""
    payload = request.get_json(silent=True) or {}
    filename = payload.get("file", "")
    job_no = payload.get("jobNo")
    
    if not job_no:
        return jsonify({"error": "jobNo is required"}), 400
    
    try:
        excel_file = _resolve_excel_file(filename)
        success = _mark_job_applied(excel_file, job_no)
        if success:
            return jsonify({
                "status": "ok", 
                "message": f"Job {job_no} marked as applied",
                "file": filename,
                "jobNo": job_no,
                "appliedAt": datetime.now().isoformat()
            }), 200
        else:
            return jsonify({"error": f"Job {job_no} not found"}), 404
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.post("/api/apply-all")
def apply_all_jobs():
    """Mark ALL jobs in a file as applied with current timestamp."""
    payload = request.get_json(silent=True) or {}
    filename = payload.get("file", "")
    
    if not filename:
        return jsonify({"error": "file is required"}), 400
    
    try:
        excel_file = _resolve_excel_file(filename)
        result = _mark_all_jobs_applied(excel_file)
        
        if result["success"]:
            return jsonify({
                "status": "ok",
                "message": f"All {result['appliedCount']} jobs marked as applied",
                "file": filename,
                "appliedCount": result["appliedCount"],
                "timestamp": result["timestamp"]
            }), 200
        else:
            return jsonify({"error": result.get("error", "Failed to apply all jobs")}), 500
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


@app.get("/api/applied-summary")
def get_applied_summary():
    """Get summary of applied jobs from Excel file."""
    filename = request.args.get("file", "")
    
    try:
        excel_file = _resolve_excel_file(filename)
        workbook = load_workbook(excel_file, data_only=True)
        worksheet = workbook.active
        
        headers = [cell.value for cell in worksheet[2]]
        
        # Find status and datetime columns
        status_col = None
        datetime_col = None
        
        for idx, header in enumerate(headers):
            if header == "Apply Status":
                status_col = idx + 1
            elif header == "Applied DateTime":
                datetime_col = idx + 1
        
        applied_jobs = []
        total_jobs = 0
        applied_count = 0
        
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, values_only=True):
            if row[0] and row[1]:  # Skip empty rows
                total_jobs += 1
                if status_col and row[status_col - 1] == "Applied":
                    applied_count += 1
                    applied_jobs.append({
                        "jobNo": row[0],
                        "title": row[1],
                        "company": row[2] if len(row) > 2 else "",
                        "appliedAt": row[datetime_col - 1] if datetime_col and datetime_col <= len(row) else "",
                    })
        
        return jsonify({
            "file": filename,
            "total": total_jobs,
            "applied": applied_count,
            "notApplied": total_jobs - applied_count,
            "appliedJobs": applied_jobs[:50]  # Return latest 50
        }), 200
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400


if __name__ == "__main__":
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "false").lower() in {"1", "true", "yes"}
    LOGGER.info("Starting API server on %s:%s debug=%s data_dir=%s", host, port, debug, DATA_DIR)
    app.run(host=host, port=port, debug=debug)
