"""
Application Tracker — Update job application status in Excel
=============================================================
Tracks which jobs have been applied to with:
- Application timestamp
- Resume/Cover letter version used
- Application status
- Notes/Follow-up info
"""

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os


class ApplicationTracker:
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.STATUS_COLORS = {
            "Not Applied": "FFFFFF",      # White
            "Draft Ready": "FFF2CC",      # Light yellow
            "Applied": "C6EFCE",          # Light green
            "Skipped": "FFC7CE",          # Light red
            "Failed": "FF0000",           # Red
            "Interview": "92D050",        # Bright green
            "Rejected": "FF0000",         # Red
            "Offer": "00B050",            # Dark green
        }

    def add_tracking_columns(self):
        """Add application tracking columns to existing Excel if not present."""
        if not os.path.exists(self.excel_file):
            print(f"❌ File {self.excel_file} not found.")
            return False
        
        wb = load_workbook(self.excel_file)
        ws = wb.active
        
        # Check if tracking columns already exist
        last_col = ws.max_column
        current_headers = [cell.value for cell in ws[2]]
        
        tracking_columns = [
            "Apply Status", "Applied DateTime", "Resume Version",
            "Cover Letter", "Application Method", "Recruiter", "Follow-up Date"
        ]
        
        # Add missing columns
        start_col = last_col + 1
        for idx, col_name in enumerate(tracking_columns):
            if col_name not in current_headers:
                col_letter = self._col_num_to_letter(start_col + idx)
                cell = ws.cell(row=2, column=start_col + idx, value=col_name)
                cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2E75B6")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(self.excel_file)
        print(f"✅ Added tracking columns to {self.excel_file}")
        return True

    def _col_num_to_letter(self, n: int) -> str:
        """Convert column number to letter."""
        result = ""
        while n > 0:
            n -= 1
            result = chr(65 + n % 26) + result
            n //= 26
        return result

    def mark_applied(self, job_no: int, resume_version: str = "Default", 
                     cover_letter_version: str = "Default", 
                     app_method: str = "External Site", 
                     recruiter: str = "", notes: str = "") -> bool:
        """Mark a job as applied."""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Find row with matching job No.
            target_row = None
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                if row[0].value == job_no:
                    target_row = row[0].row
                    break
            
            if not target_row:
                print(f"⚠️  Job No. {job_no} not found in spreadsheet")
                return False
            
            # Find tracking column indices
            header_row = ws[2]
            tracking_indices = {}
            for idx, cell in enumerate(header_row):
                if cell.value == "Apply Status":
                    tracking_indices["status"] = idx + 1
                elif cell.value == "Applied DateTime":
                    tracking_indices["datetime"] = idx + 1
                elif cell.value == "Resume Version":
                    tracking_indices["resume"] = idx + 1
                elif cell.value == "Cover Letter":
                    tracking_indices["cover_letter"] = idx + 1
                elif cell.value == "Application Method":
                    tracking_indices["method"] = idx + 1
                elif cell.value == "Recruiter":
                    tracking_indices["recruiter"] = idx + 1
                elif cell.value == "Follow-up Date":
                    tracking_indices["followup"] = idx + 1
            
            # Update cells
            if "status" in tracking_indices:
                cell = ws.cell(row=target_row, column=tracking_indices["status"])
                cell.value = "Applied"
                cell.fill = PatternFill("solid", fgColor=self.STATUS_COLORS["Applied"])
            
            if "datetime" in tracking_indices:
                ws.cell(row=target_row, column=tracking_indices["datetime"]).value = datetime.now().strftime("%Y-%m-%d %H:%M")
            
            if "resume" in tracking_indices:
                ws.cell(row=target_row, column=tracking_indices["resume"]).value = resume_version
            
            if "cover_letter" in tracking_indices:
                ws.cell(row=target_row, column=tracking_indices["cover_letter"]).value = cover_letter_version
            
            if "method" in tracking_indices:
                ws.cell(row=target_row, column=tracking_indices["method"]).value = app_method
            
            if "recruiter" in tracking_indices:
                ws.cell(row=target_row, column=tracking_indices["recruiter"]).value = recruiter
            
            wb.save(self.excel_file)
            return True
        
        except Exception as e:
            print(f"❌ Error updating tracker: {e}")
            return False

    def get_application_stats(self) -> dict:
        """Get application statistics."""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            stats = {
                "total": 0,
                "applied": 0,
                "not_applied": 0,
                "skeleton": 0,
                "failed": 0,
            }
            
            # Find tracking column
            header_row = ws[2]
            status_col = None
            for idx, cell in enumerate(header_row):
                if cell.value == "Apply Status":
                    status_col = idx + 1
                    break
            
            if status_col:
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                    status = row[status_col - 1].value or "Not Applied"
                    stats["total"] += 1
                    if status == "Applied":
                        stats["applied"] += 1
                    elif status == "Draft Ready":
                        stats["skeleton"] += 1
                    elif status == "Failed":
                        stats["failed"] += 1
                    else:
                        stats["not_applied"] += 1
            
            return stats
        
        except Exception as e:
            print(f"❌ Error getting stats: {e}")
            return {}
